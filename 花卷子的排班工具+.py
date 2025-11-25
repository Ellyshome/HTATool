# %%
try:
    import traceback
    from re import findall ,sub
    from openpyxl import load_workbook   #打开xlsx文件，返回句柄
    from openpyxl.utils import get_column_letter  #数字转化为对应列
    from copy import copy
    from os import system ,path, listdir
    from sys import argv
    from time import time,ctime
    system('mode con: cols=200 lines=40')
    base=''
except:
    print(traceback.format_exc())
    input(f'引入包时出现问题： {Exception}')



# %%
def cprint(value,ys=''):
    if ys == 'G' :print('\033[32m' +value +'\033[0m')
    elif ys == 'Y' :print('\033[33m' +value +'\033[0m')
    elif ys == 0 :print('\033[0m' +value)
    else:print('\033[91m' +value +'\033[0m')


def daemon(dee):
    def wrapper(*arg ,**karg):
        try:
            return dee(*arg ,**karg)
        except FileNotFoundError:
            cprint(f'\ndaemon:未找到指定文件,检查文件名是否正确; \n报错模块:<{dee.__name__}> ;')
            return 0
        except Exception as e:
            cprint(f'\ndaemon 反馈:遇到Exception问题; \n模块:<{dee.__name__}> ;\n错误类型:{type(e)} ;\n错误反馈: {e} ;')
            print(traceback.format_exc())
            return 0
    return wrapper
        

# %%
#人员类
class doctor:
    def __init__(self,name,section,task):
        self.cell = name
        if findall('[^\u4e00-\u9fff]',self.cell.value):
            self.name = name.value.split(findall('[^\u4e00-\u9fff]',name.value)[0])[0]
        else :
            self.name = name.value
        if len(self.name)>4 or '皮' in self.name:
            cprint(f'在表<{section}>中发现疑似非法姓名： <{self.name}>,并丢弃')
            self.section = '错误'
        else:
            self.section = section

        self.task = task
        self.cell = name
        self.row = name.row
        self.col = name.column
        self.cell_t = ''

# %%
#在分表中根据字体颜色找到所有医生。
def get_doctors(sheet): 

    path = sheet.title
    doctor_class=[]
    color=sheet['A3'].font.color #以A3的颜色为基准

    for row in range(1, sheet.max_row + 1):
        cell = sheet[f'A{row}']
        if not isinstance(cell.value, str):continue
        if cell.font.color != color:continue
        doctor_class.append(doctor(cell,path,row))
    new_doctor_class = [i for i in doctor_class if i.section!='错误']
    return new_doctor_class 


# %%
def lookfor(sheet,name,col=1): #在总表姓名栏中找对应字符（找人）
    cells = []
    for row in range(2,sheet.max_row+1):
        cell=sheet[f'{get_column_letter(col+1)}{row}']
        if cell.value == None:continue
        if len(str(cell.value))>18:
            print('|',end = '')
            continue 
        if cell.value == None or cell.value == '':
            #print('-字数问题',end = '')
            continue 
        if '皮' in copy(cell.value) or len(copy(cell.value))>10:
            #print('-字数问题',end = '')
            continue
        if name in cell.value:
            cells.append(cell)

    if len(cells) == 1:
        return cells[0]
    elif len(cells)>1:
        cprint(f'lookfor函数发现:{name}目标多于一个')
    else:
        return 0

# %%
def getmerg(cell):  #获取指定cell的 合并单元格状态
    sheet = cell.parent
    col = cell.column
    row = cell.row
    merge_cells = sheet.merged_cells.ranges
    merged = [cell for cell in merge_cells if cell.min_row <= row <= cell.max_row]
    for merged_row in merged:
        if merged_row.min_col == col :return 1      #是合并单元格，且是主单元格
        if merged_row.min_col < col <=merged_row.max_col :return 2  #是合并单元格，不是主单元格
    return 0    #不是合并单元格

# %%
# 获取指定行的合并单元格状态
def get_merged_cells(sheet,row):
    merged_cells = []
    for merged_cell in sheet.merged_cells:
        if merged_cell.min_row <= row <= merged_cell.max_row:
            merged_cells.append(merged_cell)
    return merged_cells

    # 解除所有合并单元格
def unmerge(cell):
    sheet = cell.parent
    row = cell.row
    merge_cells = sheet.merged_cells.ranges
    specified_row_merge_cells = [cell for cell in merge_cells if cell.min_row <= row <= cell.max_row]
    for merge_cell in specified_row_merge_cells:
        sheet.unmerge_cells(str(merge_cell))


    # 复制单元格合并信息  
def merge_cells_in_row(doctor,compair_flag):
    if compair_flag == 1:   #以分表覆盖总表
        unmerge(doctor.cell_t)
        sheet_s = doctor.cell.parent
        row_s = doctor.cell.row
        sheet_t = doctor.cell_t.parent
        row_t = doctor.cell_t.row
        offset = doctor.cell_t.column - doctor.cell.column
    else :      #以总表覆盖分表
        unmerge(doctor.cell)
        sheet_s = doctor.cell_t.parent
        row_s = doctor.cell_t.row
        sheet_t = doctor.cell.parent
        row_t = doctor.cell.row
        offset = doctor.cell.column - doctor.cell_t.column
    merged_cells = get_merged_cells(sheet_s,row_s)  #获取目标行单元格合并信息

    for merged_cell in merged_cells:
        start_column = merged_cell.min_col + offset
        end_column = merged_cell.max_col + offset
        row = row_t
        #print(f'合并单元格的三个参数{row}, {start_column},{end_column}')
        sheet_t.merge_cells(start_column=start_column, start_row=row, end_column=end_column, end_row=row)

# %%
def compair_cell(s_cell,t_cell): #对比 单元格
    tt = sub(r'\s+','',str(copy(t_cell.value)))
    ss = sub(r'\s+','',str(copy(s_cell.value)))

    if tt.casefold() != ss.casefold():
        cprint(f'-->\t-表 {t_cell} :[{tt}],与表  {s_cell}:[{ss}]  不一致','Y')
        return 0
    return 1

# %%
def copy_style(t_cell,s_cell):
    try:
        t_cell.font = copy(s_cell.font)         #字体（包括颜色）
        t_cell.border = copy(s_cell.border)     #边线样式
        t_cell.fill = copy(s_cell.fill)         #底色
        t_cell.alignment = copy(s_cell.alignment)   #对其方式
        if s_cell.has_style:t_cell._style = copy(s_cell._style)
        return 1
    except:
        cprint('单元格 样式 复制时出现问题')
        return 0

# %%
def copy_cell(t_cell,s_cell):  #复制 单个单元格
    try:
        t_cell.value = s_cell.value     #值复制
        #print(f'{get_column_letter(t_cell.column)}',end='')
        if not compair_cell(t_cell,s_cell): cprint('自检发现问题')
    except AttributeError:
        if not compair_cell(t_cell,s_cell): cprint('None自检发现问题')
            #print(f'{get_column_letter(t_cell.column)}',end='')
    except:
        cprint('单元格 值 复制时出现问题')
    finally:
        print('|',end='')
        #return 1 if compair_cell(t_cell,s_cell) else 0

# %%

def match(doctors, sheet_t):
    pop_doctor=[]
    for i in doctors:
        cell_t=lookfor(sheet_t,i.name)
        if cell_t == 0:
            cprint(f'{i.name}--不在总表内')
            continue 
        i.cell_t = lookfor(sheet_t,i.name)
        pop_doctor.append(i)
    cprint(f'配对成功,共{len(pop_doctor)}人','G')
    return pop_doctor

def change_sheet_s(sheets,sheet,compair_flag=0):   #任务分支  parameter：sheet集合,sheet序号
    print('-'*35)
    doctors= get_doctors(sheets[sheet])
    cprint(f'在 {sheets[sheet].title} 搜索到医生{len(doctors)}名','G')
    pop_doctor=match(doctors,sheets[0])

    
    for doctor in pop_doctor:
        print (doctor.name,end='-\t')
        if compair_flag != 0 :merge_cells_in_row(doctor,compair_flag) #不是对比则进行单元格合并统一
        for day in range(1,15):
            sheet_s = doctor.cell.parent
            row_s = doctor.cell.row
            col_s = get_column_letter(doctor.cell.column+day)

            sheet_t = doctor.cell_t.parent
            row_t = doctor.cell_t.row
            col_t = get_column_letter(doctor.cell_t.column+day)
            cell_s=sheet_s[f'{col_s}{row_s}']
            cell_t=sheet_t[f'{col_t}{row_t}']

            if compair_flag == 2: cell_t,cell_s = cell_s,cell_t 
            if compair_flag == 0:compair_cell(cell_t,cell_s)
            else : 
                copy_style(cell_t,cell_s)
                if copy_cell(cell_t,cell_s) == 0:cprint('X')
        print('')


# %%
def delflag(sheet,rown):  #用/区分上下午
    for day in range(3,17):
        for row in range(2,rown):

            day_l=get_column_letter(day)
            cell =  sheet[f'{day_l}{row}']
            if not isinstance(cell.value ,str):continue
            if not ('/' in cell.value) :continue
            if getmerg(cell) == 0: 
                cell.value = cell.value.replace('/','')
                continue
            day_n=get_column_letter(day+1)
            sheet.unmerge_cells(f'{day_l}{row}:{day_n}{row}')
            Am,Pm ,*p= cell.value.split('/')
            #if len(p)>0:print(f'有晚班 cell:{cell}   value: {p}')  
            cell.value = Am
            cell_n = sheet[f'{day_n}{row}']
            copy_style(cell_n,cell)
            cell_n.value = Pm

# %%
def statistic(sheet_t):     #统计主、专
    zd = dict()
    rown = sheet_t.max_row
    delflag(sheet_t,rown)
    for col in range(3,17):
        doctors = []
        for row in range(2,rown):
            cell = sheet_t[f'{get_column_letter(col)}{row}']
            if getmerg(cell) == 2: cell =sheet_t[f'{get_column_letter(col-1)}{row}']
            if not isinstance(cell.value ,str):continue
            try:
                if ('主' in cell.value) or ('专' in cell.value) or ('甲病' in cell.value) or ('黄褐斑门诊' in cell.value) or ('白癜风' in cell.value)or ('痤疮' in cell.value):
                    if '激' in cell.value : continue
                    if '脱' in cell.value : continue
                    if '性' in cell.value : continue
                    if '靶' in cell.value : continue
                    if '注射' in cell.value : continue
                    if '美容' in cell.value : continue
                    if '带疱' in cell.value : continue
                    if len(cell.value) >10:
                        print(f'大于10字内容:{cell.value}')
                        continue

                    doctors.append(f'{row}-{cell.value}')
            except:
                print(f'出现问题cell:{col}-{row} : {cell.value}')
        zd[col-2]= doctors
                
    return zd

# %%
def print_t(zd):   #输出统计量
    for doctor in zd:
        #week = (column_index_from_string(doctor.cell.column)-1) //2
        week ,n= str((doctor+1)/2).split('.')
        m ='Am' if n=='0'else 'Pm'
        n = len(zd[doctor])
        cprint(f'周{week}-{m} : 合 {n} 人 -- {zd[doctor]}',n>16)

# %%
def get_handle() -> str:
    global base
    base = input('请拖入排班文件,或写入带路径的文件名,按回车继续>>')
    while 1:
        if ':' not in files:# 若没有：,则判定为没有路径,将为其增加本地路径
            try:
                dir_path = path.dirname(path.realpath(__file__)) 
                for filename in listdir(dir_path):
                    if filename.split('.')[-1] == 'xlsx':
                        file_path = path.join(dir_path, filename)
                        files.append(file_path)
            except:
                files = input('当前目录没有xlsx文件,请拖入排班文件,或写入带路径的文件名,按回车继续>>')
                continue
        if not path.isfile(files):
            files = input('输入的地址不是合法文件,请拖入排班文件,或写入带路径的文件名,按回车继续>>')
            continue
        if files.split('.')[-1] != 'xlsx':
            files = input('输入的地址不是xlsx文件,请拖入排班文件,或写入文件名,按回车继续>>')
            continue
        files = input('文件路径错误,请拖入排班文件,或写入文件名,按回车继续>>>>')
    

    lasttime = ctime(path.getmtime(base))
    cprint(f'当前打开的文件是： {base}','Y')
    cprint(f'最终修改时间是： {lasttime}','Y')
    pastime = (time()-path.getmtime(base))//60
    cprint(f'距离现在过去了{pastime} 分钟')
    cprint(f'开错了别再赖我～～','Y')
    file_handle = load_workbook(base)
    return file_handle

# %%
@daemon
def main():
    get_handle()
    while True:
        cprint ('选择功能:\n\t 0、对比表\n\t 1、改总表\n\t 2、改分表\n\t 3、统计\n\t q、退出','G')
        setnumber= input('>>>')
        if setnumber == 'q':break
        file_handle=get_handle()
        if file_handle == 0:
            cprint('需要将文件改名为base.xlsx存与本目录下,或将目标文件拖拽到本图标上')
            continue
        sheets = file_handle.worksheets
        if '-' in setnumber:
            _,sheet=setnumber.split('-')
            if int(sheet)> len(sheets) : continue
            change_sheet_s(sheets,int(sheet),compair_flag=1)

        elif setnumber=='0':    #0为对比
            for sheet in range(1,len(sheets)):
                if change_sheet_s(sheets,sheet,compair_flag=0) == 0:
                    cprint('对比出现问题')
            cprint('对比完成','G')

        elif setnumber=='1':    #1为改总表
            for sheet in range(1,len(sheets)):
                if change_sheet_s(sheets,sheet,compair_flag=1) == 0:
                    cprint('替换出现问题')
            cprint('替换完成','G')

        elif setnumber=='2':    #2为改分表
            for sheet in range(1,len(sheets)):
                if change_sheet_s(sheets,sheet,compair_flag=2) == 0:
                    cprint('反改分表问题')
            cprint('反改分表完成','G')
            
        elif setnumber=='3':    #S为对比
            print('-'*35)
            print('')
            print_t(statistic(file_handle.worksheets[0]))
        else:
            system('cls')
        try:
            if setnumber == '1' or setnumber == '2' or '-'in setnumber: 
                file_handle.save(base)
                cprint('文件保存完成','G')
            else:
                cprint(f'未保存')
        except PermissionError:
            cprint('文件保存失败,可能文件正在被打开,关闭后再试')
        if(file_handle.close):cprint('-'*35,'G')
    return 1

# %%
print('-----------------花卷子的排班工具 v1.0 --------------------')
if main() == 0:
    input('按任意键退出程序')


