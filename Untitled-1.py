# %%

from re import findall
from openpyxl.utils import get_column_letter  #数字转化为对应列
from openpyxl import load_workbook
from copy import copy
from sys import argv
import os
os.system('mode con: cols=200 lines=40')
doctor_list = {}

# %%
def split_jjb(name,cell):
    row = name.row()
    for col in (3,17):
        cell_t = cell.parent[f'{get_column_letter(col)}{row}']
        jjb = str(copy(namecell.value))

# %%
def getname(cell):  #获取人名
    namecell = cell.parent[f'B{cell.row}']
    
    cellname = str(copy(namecell.value))
    key = findall('[^\u4e00-\u9fff]',cellname)
    if '/' in key : print(f'注意有换班：{cellname}')
    if key:
        doctor_name = cellname.split(key[0])[0]
    else:
        doctor_name = cellname
    if '皮' in doctor_name:doctor_name = doctor_name.replace('皮','')
    return doctor_name

# %%
def scan(keyword,filehandle):   #计数
    global doctor_dict
    sheet = load_workbook(filehandle).worksheets[0]
    cont = 0
    for row in range(1,sheet.max_row+1):
        for col in range(3,17):
            cell = sheet[f'{get_column_letter(col)}{row}']
            cellvalue = str(copy(cell.value))
            if cellvalue == None: continue
            if not (keyword in cellvalue): continue
            doctor_name = getname(cell)
            if doctor_name == 0:continue    #如果有换班
            if not (doctor_name in doctor_dict):
                doctor_dict[doctor_name] = []
            doctor_dict[doctor_name].append(cell)
            cont += 1
    return cont

# %%
def get_name_for_list(filehandle):
    global doctor_list
    sheet = load_workbook(filehandle).worksheets[0]
    for col in range(1,sheet.max_column+1):
        doctor_list[col] = []
        for row in range(2,sheet.max_row+1):
            cell = sheet[f'{get_column_letter(col)}{row}']
            cellvalue = str(copy(cell.value))
            if cellvalue == None: break
            doctor_list[col].append(cellvalue)

# %%
def getfile() -> str:
    files = input('请拖入排班文件,或写入带路径的文件名，按回车继续>>')
    while 1:
        if ':' not in files:# 若没有：，则判定为没有路径，将为其增加本地路径
            try:
                dir_path = os.path.dirname(os.path.realpath(__file__)) 
                for filename in os.listdir(dir_path):
                    if filename.split('.')[-1] == 'xlsx':
                        file_path = os.path.join(dir_path, filename)
                        files.append(file_path)
            except:
                files = input('当前目录没有xlsx文件，请拖入排班文件,或写入带路径的文件名，按回车继续>>')
                continue
        if not os.path.isfile(files):
            files = input('输入的地址不是合法文件，请拖入排班文件,或写入带路径的文件名，按回车继续>>')
            continue
        if files.split('.')[-1] != 'xlsx':
            files = input('输入的地址不是xlsx文件，请拖入排班文件,或写入文件名，按回车继续>>')
            continue
        files = input('文件路径错误，请拖入排班文件,或写入文件名，按回车继续>>>>')
    return  files

# %%
print('---------人数计算---------------')
files = []
if len(argv)>1 and argv[1].split('.')[-1] == 'xlsx':    #可以接收拖入文件
    for i in argv:
        if i.split('.')[-1] == 'xlsx':files.append(i)
        if i.split('.')[-2] == '三院名单':doctor_list[0].append(i)
else: 
    dir_path = getfile()    
    #files.append("1.xlsx")      #ipynb中测试使用

sum = len(files)
print(f'{sum}  valid files found')
for i in files:print(i)

while 1:
    global doctor_dict
    doctor_dict = {}   #key=人名，value=[cell集合]
    keyword = input('\n输入要统计的词>>>')
    if keyword == '':break
    for i in files:
        print(f'字段 【{keyword}】 在文件{i}中，匹配了 {scan(keyword,i)} 次')
    print('-'*50)
    for key,value in doctor_dict.items():
        word = str(f'{key}\t{len(value)}次:')
        words = word.replace(' ','-')
        print (f'{words} {value}')


