from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import os
import time
import shutil
import openpyxl
from openpyxl.styles import Font
import logging

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def setup_download_directory():
    """设置下载目录并确保其为空"""
    download_dir = os.path.join(os.getcwd(), 'test_downloads')
    if os.path.exists(download_dir):
        shutil.rmtree(download_dir)
    os.makedirs(download_dir)
    return download_dir

def setup_webdriver(download_dir):
    """配置并启动WebDriver"""
    chrome_options = Options()
    chrome_options.add_argument('--headless')  # 无头模式运行
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--window-size=1920,1080')
    
    # 设置下载目录
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    chrome_options.add_experimental_option("prefs", prefs)
    
    # 启动WebDriver
    driver = webdriver.Chrome(options=chrome_options)
    return driver

def upload_file(driver, file_path):
    """上传文件到网页应用"""
    logger.info(f"正在上传文件: {file_path}")
    # 找到文件输入元素并发送文件路径
    file_input = driver.find_element(By.ID, 'fileInput')
    file_input.send_keys(file_path)
    logger.info("文件上传完成")

def execute_processing(driver):
    """执行表格处理操作"""
    logger.info("开始执行表格处理操作")
    
    # 假设页面上有处理按钮，点击执行处理
    try:
        # 等待文件上传完成后界面更新
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'fileInfo'))
        )
        
        # 尝试点击可能的处理按钮（根据页面实际按钮ID调整）
        process_buttons = driver.find_elements(By.CSS_SELECTOR, 'button')
        for button in process_buttons:
            if '处理' in button.text or '运行' in button.text:
                button.click()
                logger.info(f"点击了处理按钮: {button.text}")
                break
        
        # 等待处理完成（假设处理过程中有某种状态指示）
        time.sleep(5)  # 根据实际处理时间调整
        logger.info("表格处理操作执行完成")
    except Exception as e:
        logger.error(f"执行处理操作时出错: {str(e)}")
        raise

def download_result(driver):
    """下载处理结果文件"""
    logger.info("开始下载结果文件")
    
    try:
        # 尝试点击可能的下载按钮
        download_buttons = driver.find_elements(By.CSS_SELECTOR, 'button')
        for button in download_buttons:
            if '下载' in button.text or '保存' in button.text:
                button.click()
                logger.info(f"点击了下载按钮: {button.text}")
                break
        
        # 等待下载完成
        time.sleep(5)  # 根据实际下载时间调整
        logger.info("结果文件下载完成")
    except Exception as e:
        logger.error(f"下载结果文件时出错: {str(e)}")
        raise

def get_downloaded_file(download_dir):
    """获取下载目录中最新的文件"""
    files = os.listdir(download_dir)
    if not files:
        return None
    
    # 获取最新下载的文件
    files.sort(key=lambda x: os.path.getmtime(os.path.join(download_dir, x)))
    return os.path.join(download_dir, files[-1])

def verify_color_preservation(original_file, result_file):
    """验证结果文件是否保留了原始文件的颜色信息"""
    logger.info(f"验证颜色保留情况: 原始文件={original_file}, 结果文件={result_file}")
    
    try:
        # 打开原始文件和结果文件
        original_wb = openpyxl.load_workbook(original_file)
        result_wb = openpyxl.load_workbook(result_file)
        
        # 检查每个工作表
        color_matches = 0
        color_mismatches = 0
        total_cells_checked = 0
        
        for sheet_name in original_wb.sheetnames:
            if sheet_name in result_wb.sheetnames:
                original_sheet = original_wb[sheet_name]
                result_sheet = result_wb[sheet_name]
                
                # 检查有限数量的单元格以提高效率
                max_rows = min(50, original_sheet.max_row)
                max_cols = min(10, original_sheet.max_column)
                
                logger.info(f"检查工作表 '{sheet_name}' 中的前 {max_rows} 行和 {max_cols} 列")
                
                for row in range(1, max_rows + 1):
                    for col in range(1, max_cols + 1):
                        original_cell = original_sheet.cell(row=row, column=col)
                        result_cell = result_sheet.cell(row=row, column=col)
                        
                        # 获取字体颜色
                        original_color = None
                        result_color = None
                        
                        if original_cell.font.color:
                            original_color = original_cell.font.color.rgb
                        if result_cell.font.color:
                            result_color = result_cell.font.color.rgb
                        
                        total_cells_checked += 1
                        
                        if original_color == result_color:
                            color_matches += 1
                        else:
                            color_mismatches += 1
                            logger.warning(f"单元格 [{row},{col}] 颜色不匹配 - 原始: {original_color}, 结果: {result_color}")
        
        # 计算匹配率
        match_rate = (color_matches / total_cells_checked * 100) if total_cells_checked > 0 else 0
        
        logger.info(f"颜色验证结果: 匹配={color_matches}, 不匹配={color_mismatches}, 总检查={total_cells_checked}, 匹配率={match_rate:.2f}%")
        
        return {
            'color_matches': color_matches,
            'color_mismatches': color_mismatches,
            'total_cells_checked': total_cells_checked,
            'match_rate': match_rate
        }
    
    except Exception as e:
        logger.error(f"验证颜色保留时出错: {str(e)}")
        raise

def main():
    """主测试函数"""
    original_file = os.path.join(os.getcwd(), '1.xlsx')
    
    try:
        # 设置下载目录
        download_dir = setup_download_directory()
        
        # 配置WebDriver
        driver = setup_webdriver(download_dir)
        
        # 打开网页
        url = "http://localhost:8000/1.html"
        logger.info(f"正在打开网页: {url}")
        driver.get(url)
        
        # 上传文件
        upload_file(driver, original_file)
        
        # 处理文件
        execute_processing(driver)
        
        # 下载结果
        download_result(driver)
        
        # 获取下载的文件
        result_file = get_downloaded_file(download_dir)
        
        if result_file:
            logger.info(f"成功下载结果文件: {result_file}")
            
            # 验证颜色保留
            verification_result = verify_color_preservation(original_file, result_file)
            
            # 输出结果
            print("\n===== 测试结果摘要 =====")
            print(f"原始文件: {original_file}")
            print(f"结果文件: {result_file}")
            print(f"颜色匹配: {verification_result['color_matches']}")
            print(f"颜色不匹配: {verification_result['color_mismatches']}")
            print(f"总检查单元格: {verification_result['total_cells_checked']}")
            print(f"颜色匹配率: {verification_result['match_rate']:.2f}%")
            
            if verification_result['match_rate'] > 90:
                print("\n✅ 测试通过: 表格样式和颜色信息得到了良好保留")
            else:
                print("\n❌ 测试失败: 表格样式和颜色信息未完全保留")
        else:
            logger.error("未能找到下载的结果文件")
            print("\n❌ 测试失败: 未能找到下载的结果文件")
            
    except Exception as e:
        logger.error(f"测试过程中发生错误: {str(e)}")
        print(f"\n❌ 测试失败: {str(e)}")
        
    finally:
        # 关闭浏览器
        if 'driver' in locals():
            driver.quit()

if __name__ == "__main__":
    main()